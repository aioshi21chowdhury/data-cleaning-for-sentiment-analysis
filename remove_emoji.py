from openpyxl import load_workbook
import emoji

# Function to replace emojis with their word meanings
def replace_emojis(text):
    # Convert emojis to their descriptions
    return emoji.demojize(text)

# Load the Excel workbook
wb = load_workbook('sheet2_sheet1.xlsx')

# Get the sheet with data
sheet = wb['Sheet1']

# Iterate through each cell in the sheet
for row in sheet.iter_rows():
    for cell in row:
        if isinstance(cell.value, str):
            # Replace emojis in the cell value
            cell.value = replace_emojis(cell.value)

# Save the changes to the Excel file
wb.save('modified_emoji_file.xlsx')

print("Emojis replaced with their word meanings in Sheet1.")
